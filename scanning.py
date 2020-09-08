import csv
import openpyxl as xl;
import sys
from openpyxl.styles import Font

# This program converse the source file from CSV to XLSX. Extract rows when the prefix of particular cells matches with the Keywords.

# ===============Variables for the script=================#
targetExcel = "Test_Output.xlsx"
prefixKeyword = "CVE"
referenceName = "CVE"
referenceRow = 1

sourcecsv = 'a.csv'
targetxlsx = 'd.xlsx'


# ========================================================#

def conversion(source, target):
    wb = xl.Workbook()
    ws = wb.active

    with open(source, 'r', encoding='UTF-8') as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            ws.append(row)

    wb.save(target)


def autoextract(inputfile, outputfile):
    # opening the source excel file
    filename = inputfile
    wb1 = xl.load_workbook(filename)

    # opening the destination excel file
    filename1 = outputfile
    wb2 = xl.load_workbook(filename1)
    ws2 = wb2["Output"]

    sheets = wb1.sheetnames
    print("The sheets for processing include:", sheets)
    x = len(sheets)
    for z in range(x):
        ws1 = wb1.worksheets[z]

        # calculate total number of rows and
        # columns in the source excel file
        mr = ws1.max_row
        mc = ws1.max_column

        # Initialize a variable for checking the column position of reference value.
        referVal = 999
        for a in range(1, mc + 1):
            if ws1.cell(row=referenceRow, column=a).value == referenceName:
                referVal = a
                break

        if referVal == 999:
            print("ERROR: No Reference Name cell was found.")
            sys.exit(1)
        #print("referVal:", referVal)

        # Set header
        ws2['A1'] = "Plugin ID"
        ws2['B1'] = "CVE"
        ws2['C1'] = "CVSS"
        ws2['D1'] = "Risk"
        ws2['E1'] = "Host"
        ws2['F1'] = "Protocol"
        ws2['G1'] = "Port"
        ws2['H1'] = "Name"

        # Set font style
        ws2['A1'].font = Font(bold=True)
        ws2['B1'].font = Font(bold=True)
        ws2['C1'].font = Font(bold=True)
        ws2['D1'].font = Font(bold=True)
        ws2['E1'].font = Font(bold=True)
        ws2['F1'].font = Font(bold=True)
        ws2['G1'].font = Font(bold=True)
        ws2['H1'].font = Font(bold=True)



        # copying the cell values from source
        # excel file to destination excel file
        for i in range(1, mr + 1):
            if i > 2:
                tempText = ws1.cell(row=i, column=referVal).value

                # Only compare with Keyword when the cell is NOT empty.
                if tempText:
                    if tempText.startswith(prefixKeyword):
                        lastrow = len(ws2['A'])  # Check the last row of Column A for appending.
                        for j in range(1, mc + 1):
                            # reading cell value from source excel file
                            c = ws1.cell(row=i, column=j)

                            # writing the read value to destination excel file
                            ws2.cell(row=lastrow + 1, column=j).value = c.value

    # saving the destination excel file
    wb2.save(str(filename1))


def clearsheet(outputfile):
    wb2 = xl.load_workbook(outputfile)
    ws2 = wb2["Output"]
    for row in ws2['A1:Z999']:
        for cell in row:
            cell.value = None
    wb2.save(str(outputfile))


if __name__ == '__main__':
    try:
        conversion(sourcecsv, targetxlsx)
        sourceExcel = targetxlsx
        clearsheet(targetExcel)
        autoextract(sourceExcel, targetExcel)

        print("The program has been completed. Please check the output file:", targetExcel)
    except:
        print("Unexpected error.")
        #print("Unexpected error:", sys.exc_info()[0])
        raise
